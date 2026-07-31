#!/usr/bin/env python3
"""
gera_hv_mensal.py
Extrai base de veículos elegíveis para HV (mês anterior) via Trino.
Output: CSV em ~/Documents/HV Bases/hv_base_YYYY-MM.csv
"""

import json
import subprocess
import csv
import os
import logging
from datetime import datetime, timedelta

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

OUTPUT_DIR = os.path.expanduser("~/Documents/HV Bases")
TRINO_BRIDGE = "/Users/[YOUR_USER]/Documents/Main/Brain/Bridges/trino_mcp_bridge.py"

QUERY = """
WITH base_filtrada AS (
    SELECT 
        ad.list_id_nk,
        ad.ad_id_nk,
        ad.account_id_fk,
        veh.vehicle_tag,
        ROW_NUMBER() OVER(PARTITION BY veh.vehicle_tag ORDER BY ad.ad_id_nk DESC) as rn
    FROM 
        hive.ods.ad AS ad
    INNER JOIN 
        hive.olx_auto.listing_vehicle AS veh 
        ON ad.list_id_nk = veh.list_id_nk
    WHERE 
        ad.year = EXTRACT(YEAR FROM CURRENT_DATE - INTERVAL '1' MONTH)
        AND ad.month = EXTRACT(MONTH FROM CURRENT_DATE - INTERVAL '1' MONTH)
        AND ad.category_id_fk = 46 
        AND ad.deletion_date IS NULL 
        AND ad.ad_type_id_fk = 1
        AND veh.regdate >= 2024
        AND veh.vehicle_tag IS NOT NULL 
        AND TRIM(veh.vehicle_tag) <> ''
),
base_unica AS (
    SELECT 
        list_id_nk,
        ad_id_nk,
        account_id_fk,
        vehicle_tag
    FROM 
        base_filtrada
    WHERE 
        rn = 1
)
SELECT 
    bu.list_id_nk,
    bu.ad_id_nk,
    bu.vehicle_tag,
    acc.cpf,
    acc.account_id_pk,
    acc.account_id_nk,
    bu.account_id_fk
FROM 
    base_unica bu
INNER JOIN 
    hive.ods.account acc ON bu.account_id_fk = acc.account_id_pk
WHERE 
    acc.cpf IS NOT NULL 
    AND TRIM(acc.cpf) <> ''
    AND NOT EXISTS (
        SELECT 1 
        FROM hive.olx_vas_premium.vehicle_history_daily vhd
        WHERE CAST(bu.list_id_nk AS VARCHAR) = vhd.list_id
          AND vhd.provider_queued_status = 'finished'
          AND vhd.paid = true
          AND (
              vhd.status = 'available'
              OR (
                  vhd.status IN ('manually_excluded', 'ad_excluded')
                  AND (vhd.provider_name = 'checktudo' AND vhd.provider_response_code = '200')
              )
          )
    )
ORDER BY 
    RANDOM() 
LIMIT 10000
"""


def trino_query(sql, schema="ods"):
    """Executa query via Trino MCP bridge."""
    payload = json.dumps({
        "jsonrpc": "2.0", "id": 1,
        "method": "tools/call",
        "params": {"name": "query", "arguments": {"sql": sql, "schema": schema}}
    })
    result = subprocess.run(
        ["python3", TRINO_BRIDGE],
        input=payload, capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError(f"Trino bridge error: {result.stderr}")
    data = json.loads(result.stdout)
    if "error" in data:
        raise RuntimeError(f"Trino query error: {data['error'].get('message', data['error'])}")
    return data["result"]["content"][0]["text"]


def parse_trino_result(raw_text):
    """Converte resultado do Trino (formato bridge) em lista de dicts."""
    import ast
    
    lines = raw_text.strip().split("\n")
    if not lines:
        return [], []
    
    # Formato: "Colunas: ['col1', 'col2', ...]\nDados: [[val1, val2, ...], ...]"
    headers_line = lines[0]
    data_line = "\n".join(lines[1:])
    
    # Extrair headers
    headers_str = headers_line.replace("Colunas: ", "")
    headers = ast.literal_eval(headers_str)
    
    # Extrair dados
    data_str = data_line.replace("Dados: ", "")
    data = ast.literal_eval(data_str)
    
    rows = [dict(zip(headers, row)) for row in data]
    return headers, rows


def main():
    # Mês de referência (mês anterior)
    ref_date = datetime.now().replace(day=1) - timedelta(days=1)
    ref_label = ref_date.strftime("%Y-%m")
    
    logger.info(f"Gerando base HV para o mês de referência: {ref_label}")
    logger.info("Executando query no Trino...")
    
    raw_result = trino_query(QUERY)
    headers, rows = parse_trino_result(raw_result)
    
    if not rows:
        logger.warning("Nenhum resultado retornado pela query.")
        return
    
    logger.info(f"Total de registros: {len(rows)}")
    
    # Garantir que a pasta existe
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Salvar XLSX com colunas ajustadas
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    filename = f"hv_base_{ref_label}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HV Base"
    
    # Header
    ws.append(headers)
    
    # Dados
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    
    # Ajustar largura das colunas
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
    
    wb.save(filepath)
    
    logger.info(f"✅ Arquivo exportado: {filepath}")
    logger.info(f"Total de linhas: {len(rows)}")


if __name__ == "__main__":
    main()
